from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import Workbook, load_workbook
from threading import Thread
import tkinter as tk

def print_pro(text):
    print(text)
    text_logs.insert(tk.END, text+"\n")
    text_logs.see("end")


def goto(section,driver):
    print_pro(section)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/span/div/div/div/ul/li[1]')))
    time.sleep(4)
    num = 1
    while(True):
        check_section = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/span/div/div/div/ul/li['+str(num)+']/a')
        if(check_section.text == section):
            check_section.click()
            break
        num+=1

def open_match_page():
    game_link = ent_gameurl.get()
    driver_manual = webdriver.Chrome()
    driver_manual.get(game_link)

def create_empty_list(count):
    empty_list=[]
    for i in range(0, count):
        empty_list.append("")
    return empty_list

is_run = True
def start_press():
    print_pro("Starting ...")
    global is_run
    is_run = True

    t = Thread(target=start_checking)
    t.start()

def stop_checking():
    print_pro("Stopping ...")
    global is_run
    is_run = False

def start_checking():
    game_link = ent_gameurl.get()
    sleep_time = int(ent_intervaltime.get())

    options = Options()
    options.add_argument('--headless')
    driver = webdriver.Chrome(options=options)
    # driver = webdriver.Chrome()
    driver.get(game_link)

    #cookie
    try:
        print_pro("waiting for cookie popup")
        WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]')))
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="onetrust-accept-btn-handler"]').click()
        print_pro("accept cookie popup")
    except:
        print_pro("no cookie popup")
        pass

    # Match details for file name + create file
    try:
        goto("Match Odds",driver)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))

        for j in range(0, 15):
            caption = driver.find_element_by_xpath(
                '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
            if (caption == "Match Odds"):
                break
            time.sleep(1)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div')))


        #showing game basic data in ui
        date_or_inplay = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div').text.split(
            ",")[0]
        if ("-" in date_or_inplay):
            label_date.config(text="")
        else:
            label_date.config(text=date_or_inplay)
        ui_country = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/a').text
        label_country.config(text=ui_country)
        ui_series = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text
        label_series.config(text=ui_series)
        match_name = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text
        file_name = match_name
        label_teams.config(text=match_name)
        print_pro(file_name)

    except:
        print_pro("[NO]* Match Odds -> match details for file name")
        file_name = "output"

    wb = Workbook()
    ws = wb.active

    topic_row = open("topics.txt", "r").read().split("\n")
    ws.append(topic_row)

    wb.save(file_name + '.xlsx')
    print_pro("file created: " + file_name)

    while(True):
        if is_run == False:
            break
        #Match details
        try:
            goto("Match Odds",driver)

            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                            '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
            for j in range(0, 15):
                caption = driver.find_element_by_xpath(
                    '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
                if (caption == "Match Odds"):
                    break
                time.sleep(1)
            match_detials_list = []
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div')))

            #date
            try:
                date_or_inplay = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div').text.split(",")[0]
                if("-" in date_or_inplay):
                    match_detials_list.append("")
                else:
                    match_detials_list.append(date_or_inplay)
            except:
                match_detials_list.append("")
                pass

            match_detials_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/a').text)
            match_detials_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text)
            match_detials_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text)

            #score
            try:
                WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_elements(By.XPATH, '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/span/span')
                                   or driver.find_elements(By.XPATH, '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div[1]/ng-include/div/div/span[2]'))

                result_list = driver.find_elements(By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div[1]/ng-include/div/div/span[2]')
                if (len(result_list)>0):
                    match_detials_list.append(result_list[0].text)
                    label_score.config(text=result_list[0].text)
                else:
                    match_detials_list.append("")
            except:
                match_detials_list.append("")
                pass

            #mins
            try:
                WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_elements(By.XPATH,
                                                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div')
                                   or driver.find_elements(By.XPATH,
                                                           '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div[1]/p/span'))

                mins_list = driver.find_elements(By.XPATH,
                                                   '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div[1]/p/span')
                if (len(mins_list) > 0):
                    if(mins_list[0].text=="Finished"):
                        print_pro("Game is finished")
                        break
                    else:
                        match_detials_list.append(mins_list[0].text)
                        label_mins.config(text=mins_list[0].text)
                else:
                    match_detials_list.append("")
            except:
                match_detials_list.append("")
                pass

            #game half
            match_detials_list.append("")
            print(match_detials_list)

        except:
            print_pro("[NO]* Match Odds -> match details")
            match_detials_list = create_empty_list(7)

        if is_run == False:
            break
        #Match Odds
        try:
            goto("Match Odds",driver)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                            '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
            for j in range(0, 15):
                caption = driver.find_element_by_xpath(
                    '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
                if (caption == "Match Odds"):
                    break
                time.sleep(1)

            match_odds_list = []
            match_odds_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            match_odds_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            match_odds_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            match_odds_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            match_odds_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[3]/td[4]/button/div/span[1]').text)
            match_odds_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[3]/td[5]/button/div/span[1]').text)
            print(match_odds_list)

        except:
            print_pro("[NO]* Match Odds")
            match_odds_list = create_empty_list(6)

        if is_run == False:
            break

        # Half Time
        try:
            goto("Half Time", driver)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                            '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))

            for j in range(0, 15):
                caption = driver.find_element_by_xpath(
                    '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
                if (caption == "Half Time"):
                    break
                time.sleep(1)
            half_time_list = []
            for i in range(1, 4):
                for j in range(4, 6):
                    half_time_list.append(driver.find_element_by_xpath(
                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[' + str(
                            i) + ']/td[' + str(j) + ']/button/div/span[1]').text)
            print(half_time_list)
        except:
            print_pro("[NO]* Half Time")
            half_time_list = create_empty_list(6)

        if is_run == False:
            break

        #First Half Goals x.5
        first_half_goals_x5 = []
        for i in range(0,3):
            if is_run == False:
                break
            try:
                goto("First Half Goals "+str(i)+".5",driver)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
                for j in range(0,15):
                    caption = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text.split(" ")
                    if(len(caption)>2 and caption[3]==str(i)+".5"):
                        break
                    time.sleep(1)
                first_half_goals_x5.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
                first_half_goals_x5.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
                first_half_goals_x5.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
                first_half_goals_x5.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
                print(first_half_goals_x5)
            except:
                print_pro("[NO]* First Half Goals "+str(i)+".5")
                first_half_goals_x5+=["","","",""]

        if is_run == False:
            break

        #Half Time Score
        try:
            half_time_score_list = []
            goto("Half Time Score",driver)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
            for j in range(0,15):
                caption = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
                if(caption=="Half Time Score"):
                    break
                time.sleep(1)
            for i in range(1,11):
                for j in range(4,6):
                    half_time_score_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr['+str(i)+']/td['+str(j)+']/button/div/span[1]').text)

            print(half_time_score_list)
        except:
            print_pro("[NO]* Half Time Score")
            half_time_score_list = create_empty_list(20)

        if is_run == False:
            break

        #Over/under x.5 Goals
        over_under_x5_goals = []
        for i in range(0,9):
            if is_run == False:
                break
            try:
                goto("Over/Under "+str(i)+".5 Goals",driver)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
                for j in range(0,15):
                    caption = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text.split(" ")
                    if(caption[1]==str(i)+".5"):
                        break
                    time.sleep(1)
                over_under_x5_goals.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
                over_under_x5_goals.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
                over_under_x5_goals.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
                over_under_x5_goals.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
                print(over_under_x5_goals)
            except:
                print_pro("[NO]* Over/Under "+str(i)+".5 Goals")
                over_under_x5_goals+=["","","",""]

        if is_run == False:
            break

        #Correct Score
        try:
            correct_score_list = []
            goto("Correct Score",driver)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
            for j in range(0,15):
                caption = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
                if(caption=="Correct Score"):
                    break
                time.sleep(1)
            for i in range(1,20):
                for j in range(4,6):
                    correct_score_list.append(driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr['+str(i)+']/td['+str(j)+']/button/div/span[1]').text)

            print(correct_score_list)
        except:
            print_pro("[NO]* Correct Score")
            correct_score_list = create_empty_list(38)

        if is_run == False:
            break

        #Both teams to Score?
        try:
            both_teams_to_score_list =[]
            goto("Both teams to Score?",driver)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                            '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
            for j in range(0, 15):
                caption = driver.find_element_by_xpath(
                    '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
                if (caption == "Both teams to Score?"):
                    break
                time.sleep(1)
            for i in range(1, 3):
                for j in range(4, 6):
                    both_teams_to_score_list.append(driver.find_element_by_xpath(
                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[' + str(
                            i) + ']/td[' + str(j) + ']/button/div/span[1]').text)

            print(both_teams_to_score_list)
        except:
            print_pro("[NO]* Both teams to Score?")
            both_teams_to_score_list= create_empty_list(4)

        final_row = match_detials_list+half_time_list+first_half_goals_x5+half_time_score_list+match_odds_list+over_under_x5_goals+correct_score_list+both_teams_to_score_list
        wb = load_workbook(file_name+'.xlsx')
        ws = wb.active
        ws.append(final_row)
        wb.save(file_name+'.xlsx')
        print_pro("new data row added")

        driver.close()
        if is_run == False:
            break
        print_pro("sleeping for "+str(sleep_time)+" seconds")
        time.sleep(sleep_time)
        print_pro("woke up")
        if is_run == False:
            break

        options = Options()
        options.add_argument('--headless')
        driver = webdriver.Chrome(options=options)
        # driver = webdriver.Chrome()
        driver.get(game_link)

        # cookie
        try:
            print_pro("waiting for cookie popup")
            WebDriverWait(driver, 25).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]')))
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="onetrust-accept-btn-handler"]').click()
            print_pro("accept cookie popup")
        except:
            print_pro("no cookie popup")
            pass


# game_link = "https://www.betfair.com/exchange/plus/football/market/1.183636064"
# sleep_time = 10

window = tk.Tk()
window.geometry("700x360")
window.title("Football data scraper")
label_gameurl = tk.Label(text="Game url")
label_gameurl.place(x=10,y=10)
ent_gameurl = tk.Entry(width=102)
ent_gameurl.place(x=70,y=10)

label_intervaltime = tk.Label(text="Time interval (in seconds)")
label_intervaltime.place(x=10,y=40)
ent_intervaltime = tk.Entry(width=5)
ent_intervaltime.place(x=160,y=40)

btn_start = tk.Button(text="Start", width=25, command=start_press)
btn_start.place(x=300,y=38)
btn_stop = tk.Button(text="Stop", width=25, command=stop_checking)
btn_stop.place(x=500,y=38)

label_date = tk.Label(text="", width=30, anchor='w')
label_date.place(x=10,y=70)
label_country = tk.Label(text="", width=30, anchor='w')
label_country.place(x=10,y=90)
label_series = tk.Label(text="", width=30, anchor='w')
label_series.place(x=10,y=110)
label_teams = tk.Label(text="", width=30, anchor='w')
label_teams.place(x=10,y=130)

label_score = tk.Label(text="", width=20, anchor='w')
label_score.place(x=300,y=90)
label_mins = tk.Label(text="", width=20, anchor='w')
label_mins.place(x=300,y=110)
btn_gameurl = tk.Button(text="Click here to open Betfair webpage ", width=30,command=open_match_page)
btn_gameurl.place(x=300,y=130)

text_logs = tk.Text(width=84, height=10)
text_logs.place(x=10, y=180)
window.mainloop()