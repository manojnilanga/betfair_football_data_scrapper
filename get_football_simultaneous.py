from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import Workbook, load_workbook
from threading import Thread
import tkinter as tk

def rClicker(e):
    try:
        def rClick_Copy(e, apnd=0):
            e.widget.event_generate('<Control-c>')
        def rClick_Cut(e):
            e.widget.event_generate('<Control-x>')
        def rClick_Paste(e):
            e.widget.event_generate('<Control-v>')
        e.widget.focus()
        nclst=[
               (' Cut', lambda e=e: rClick_Cut(e)),
               (' Copy', lambda e=e: rClick_Copy(e)),
               (' Paste', lambda e=e: rClick_Paste(e)),
               ]
        rmenu = tk.Menu(None, tearoff=0, takefocus=0)
        for (txt, cmd) in nclst:
            rmenu.add_command(label=txt, command=cmd)
        rmenu.tk_popup(e.x_root+40, e.y_root+10,entry="0")
    except tk.TclError:
        print (' - rClick menu, something wrong')
        pass
    return "break"


def rClickbinder(r):
    try:
        for b in [ 'Text', 'Entry', 'Listbox', 'Label']: #
            r.bind_class(b, sequence='<Button-3>',
                         func=rClicker, add='')
    except tk.TclError:
        print (' - rClickbinder, something wrong')
        pass

def print_pro(text):
    print(text)
    text_logs.insert(tk.END, text+"\n")
    text_logs.see("end")


def goto(section,driver):
    print_pro(section)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/span/div/div/div/ul/li[1]')))
    time.sleep(4)
    num = 1
    while(True):
        check_section = driver.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/span/div/div/div/ul/li['+str(num)+']/a')
        if(check_section.text == section):
            check_section.click()
            break
        num+=1

def open_match_page():
    game_link = ent_gameurl.get()
    driver_manual = webdriver.Chrome()
    driver_manual.get(game_link)

def create_empty_list(count):
    empty_list=[]
    for i in range(0, count):
        empty_list.append("")
    return empty_list

is_run = True
def start_press():
    print_pro("Starting ...")
    global is_run
    is_run = True

    t = Thread(target=start_checking)
    t.start()

def stop_checking():
    print_pro("Stopping ...")
    global is_run
    is_run = False

def accept_cookie(driver):
    try:
        print_pro("waiting for cookie popup")
        WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]')))
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="onetrust-accept-btn-handler"]').click()
        print_pro("accept cookie popup")
    except:
        print_pro("no cookie popup")
        pass

def create_driver(game_link):
    options = Options()
    options.add_argument('--headless')
    driver = webdriver.Chrome(options=options)
    # driver = webdriver.Chrome()
    driver.get(game_link)
    # cookie
    accept_cookie(driver)
    return driver

def get_basic_data_and_create_file(driver):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2')))
    try:
        for j in range(0, 15):
            caption = driver.find_element_by_xpath(
                '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/h2').text
            if (caption == "Match Odds"):
                break
            time.sleep(1)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div')))


        #showing game basic data in ui
        date_or_inplay = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div').text.split(
            ",")[0]
        if ("-" in date_or_inplay):
            label_date.config(text="")
        else:
            label_date.config(text=date_or_inplay)
        ui_country = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/a').text
        label_country.config(text=ui_country)
        ui_series = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text
        label_series.config(text=ui_series)
        match_name = driver.find_element_by_xpath(
            '//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text
        file_name = match_name
        label_teams.config(text=match_name)
        print_pro(file_name)

    except:
        print_pro("[NO]* Match Odds -> match details for file name")
        file_name = "output"

    wb = Workbook()
    ws = wb.active

    topic_row = open("topics.txt", "r").read().split("\n")
    ws.append(topic_row)

    wb.save(file_name + '.xlsx')
    print_pro("file created: " + file_name)
    return file_name

all_game_links=[]
def get_game_links():
    driver = create_driver(ent_gameurl.get())
    accept_cookie(driver)
    try:
        print_pro("Getting game link of Match Odds")
        goto("Match Odds", driver)
        all_game_links.append(driver.current_url)
        file_name = get_basic_data_and_create_file(driver)
    except:
        print_pro("***** Match Odds page not found please restart *****")

    all_games=["Half Time","First Half Goals 0.5","First Half Goals 1.5","First Half Goals 2.5","Half Time Score","Over/Under 0.5 Goals","Over/Under 1.5 Goals","Over/Under 2.5 Goals","Over/Under 3.5 Goals","Over/Under 4.5 Goals","Over/Under 5.5 Goals","Over/Under 6.5 Goals","Correct Score","Both teams to Score?"]

    for i in range(0, len(all_games)):
        try:
            print_pro("Getting game link of "+all_games[i])
            goto(all_games[i], driver)
            all_game_links.append(driver.current_url)
        except:
            all_game_links.append("")

    driver.close()
    print(all_game_links)
    return file_name


def start_checking():
    sleep_time = int(ent_intervaltime.get())
    file_name = get_game_links()

    #initial loading of all browsers
    print_pro("initial loading of all game links")
    if(all_game_links[0]!=""):
        print_pro("Loading a unique browser for Match Odds")
        driver_matchodds = create_driver(all_game_links[0])
    if (all_game_links[1] != ""):
        print_pro("Loading a unique browser for Half Time")
        driver_halftime = create_driver(all_game_links[1])
    if (all_game_links[2] != ""):
        print_pro("Loading a unique browser for First Half Goals 0.5")
        driver_firsthalfgoals05 = create_driver(all_game_links[2])
    if (all_game_links[3] != ""):
        print_pro("Loading a unique browser for First Half Goals 1.5")
        driver_firsthalfgoals15 = create_driver(all_game_links[3])
    if (all_game_links[4] != ""):
        print_pro("Loading a unique browser for First Half Goals 2.5")
        driver_firsthalfgoals25 = create_driver(all_game_links[4])
    if (all_game_links[5] != ""):
        print_pro("Loading a unique browser for Half Time Score")
        driver_halftimescore = create_driver(all_game_links[5])
    if (all_game_links[6] != ""):
        print_pro("Loading a unique browser for Over/Under 0.5 Goals")
        driver_overunder05 = create_driver(all_game_links[6])
    if (all_game_links[7] != ""):
        print_pro("Loading a unique browser for Over/Under 1.5 Goals")
        driver_overunder15 = create_driver(all_game_links[7])
    if (all_game_links[8] != ""):
        print_pro("Loading a unique browser for Over/Under 2.5 Goals")
        driver_overunder25 = create_driver(all_game_links[8])
    if (all_game_links[9] != ""):
        print_pro("Loading a unique browser for Over/Under 3.5 Goals")
        driver_overunder35 = create_driver(all_game_links[9])
    if (all_game_links[10] != ""):
        print_pro("Loading a unique browser for Over/Under 4.5 Goals")
        driver_overunder45 = create_driver(all_game_links[10])
    if (all_game_links[11] != ""):
        print_pro("Loading a unique browser for Over/Under 5.5 Goals")
        driver_overunder55 = create_driver(all_game_links[11])
    if (all_game_links[12] != ""):
        print_pro("Loading a unique browser for Over/Under 6.5 Goals")
        driver_overunder65 = create_driver(all_game_links[12])
    if (all_game_links[13] != ""):
        print_pro("Loading a unique browser for Correct Score")
        driver_correctscore = create_driver(all_game_links[13])
    if (all_game_links[14] != ""):
        print_pro("Loading a unique browser for Both teams to Score?")
        driver_bothteamstoscore = create_driver(all_game_links[14])


    while(True):
        if is_run == False:
            break
        print_pro("Refreshing all the browsers ...")
        try:
            driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on matchodds")
        except:
            pass
        try:
            driver_halftime.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on halftime")
        except:
            pass
        try:
            driver_firsthalfgoals05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on firsthalfgoals05")
        except:
            pass
        try:
            driver_firsthalfgoals15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on firsthalfgoals15")
        except:
            pass
        try:
            driver_firsthalfgoals25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on firsthalfgoals25")
        except:
            pass
        try:
            driver_halftimescore.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on halftimescore")
        except:
            pass
        try:
            driver_overunder05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder05")
        except:
            pass
        try:
            driver_overunder15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder15")
        except:
            pass
        try:
            driver_overunder25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder25")
        except:
            pass
        try:
            driver_overunder35.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder35")
        except:
            pass
        try:
            driver_overunder45.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder45")
        except:
            pass
        try:
            driver_overunder55.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder55")
        except:
            pass
        try:
            driver_overunder65.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on overunder65")
        except:
            pass
        try:
            driver_correctscore.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on correctscore")
        except:
            pass
        try:
            driver_bothteamstoscore.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[1]/bf-marketview-header-wrapper/div/div/mv-header/div/div/div[2]/div/button').click()
            print_pro("Clicked refresh on bothteamstoscore")
        except:
            pass

        if is_run == False:
            break
        print_pro("Getting data from all the browsers ...")
        #Match details
        try:
            match_detials_list=[]
            #date
            try:
                date_or_inplay = driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div/ng-include/div').text.split(",")[0]
                if("-" in date_or_inplay):
                    match_detials_list.append("")
                else:
                    match_detials_list.append(date_or_inplay)
            except:
                match_detials_list.append("")
                pass

            match_detials_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/a').text)
            match_detials_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text)
            match_detials_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/div/div[4]/div/bf-navigation-lhm/div/div/ng-include/div/div/ul/li[2]/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/tree-section/ul/li/a').text)

            #score
            try:
                result_list = driver_matchodds.find_elements(By.XPATH,'//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div[1]/ng-include/div/div/span[2]')
                if (len(result_list)>0):
                    match_detials_list.append(result_list[0].text)
                    label_score.config(text=result_list[0].text)
                else:
                    match_detials_list.append("")
            except:
                match_detials_list.append("")
                pass

            #mins
            try:
                mins_list = driver_matchodds.find_elements(By.XPATH,
                                                   '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[1]/div/bf-sports-header/div/div/div/ng-include/div[1]/p/span')
                if (len(mins_list) > 0):
                    if(mins_list[0].text=="Finished"):
                        print_pro("Game is finished")
                        break
                    else:
                        match_detials_list.append(mins_list[0].text)
                        label_mins.config(text=mins_list[0].text)
                else:
                    match_detials_list.append("")
            except:
                match_detials_list.append("")
                pass

            #game half
            match_detials_list.append("")
            print(match_detials_list)

        except:
            print_pro("[NO]* Match Odds -> match details")
            match_detials_list = create_empty_list(7)

        if is_run == False:
            break
        #Match Odds
        try:
            match_odds_list = []
            match_odds_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            match_odds_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            match_odds_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            match_odds_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            match_odds_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[3]/td[4]/button/div/span[1]').text)
            match_odds_list.append(driver_matchodds.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[3]/td[5]/button/div/span[1]').text)
            print(match_odds_list)

        except:
            print_pro("[NO]* Match Odds")
            match_odds_list = create_empty_list(6)

        if is_run == False:
            break

        # Half Time
        try:
            half_time_list = []
            for i in range(1, 4):
                for j in range(4, 6):
                    half_time_list.append(driver_halftime.find_element_by_xpath(
                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[' + str(
                            i) + ']/td[' + str(j) + ']/button/div/span[1]').text)
            print(half_time_list)
        except:
            print_pro("[NO]* Half Time")
            half_time_list = create_empty_list(6)

        if is_run == False:
            break

        #First Half Goals x.5
        first_half_goals_x5 = []


        if is_run == False:
            break
        # 0.5
        try:
            first_half_goals_x5.append(driver_firsthalfgoals05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(first_half_goals_x5)
        except:
            print_pro("[NO]* First Half Goals 0.5")
            first_half_goals_x5+=["","","",""]
        # 1.5
        try:
            first_half_goals_x5.append(driver_firsthalfgoals15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(first_half_goals_x5)
        except:
            print_pro("[NO]* First Half Goals 1.5")
            first_half_goals_x5+=["","","",""]
        # 2.5
        try:
            first_half_goals_x5.append(driver_firsthalfgoals25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            first_half_goals_x5.append(driver_firsthalfgoals25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(first_half_goals_x5)
        except:
            print_pro("[NO]* First Half Goals 2.5")
            first_half_goals_x5+=["","","",""]

        if is_run == False:
            break

        #Half Time Score
        try:
            half_time_score_list = []
            for i in range(1,11):
                for j in range(4,6):
                    half_time_score_list.append(driver_halftimescore.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr['+str(i)+']/td['+str(j)+']/button/div/span[1]').text)
            print(half_time_score_list)
        except:
            print_pro("[NO]* Half Time Score")
            half_time_score_list = create_empty_list(20)

        if is_run == False:
            break

        #Over/under x.5 Goals
        over_under_x5_goals = []
        if is_run == False:
            break
        #0.5
        try:
            over_under_x5_goals.append(driver_overunder05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder05.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 0.5 Goals")
            over_under_x5_goals+=["","","",""]
        #1.5
        try:
            over_under_x5_goals.append(driver_overunder15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder15.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 1.5 Goals")
            over_under_x5_goals+=["","","",""]
        #2.5
        try:
            over_under_x5_goals.append(driver_overunder25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder25.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 2.5 Goals")
            over_under_x5_goals+=["","","",""]
        #3.5
        try:
            over_under_x5_goals.append(driver_overunder35.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder35.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder35.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder35.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 3.5 Goals")
            over_under_x5_goals+=["","","",""]
        #4.5
        try:
            over_under_x5_goals.append(driver_overunder45.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder45.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder45.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder45.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 4.5 Goals")
            over_under_x5_goals+=["","","",""]
        #5.5
        try:
            over_under_x5_goals.append(driver_overunder55.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder55.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder55.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder55.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 5.5 Goals")
            over_under_x5_goals+=["","","",""]
        #6.5
        try:
            over_under_x5_goals.append(driver_overunder65.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder65.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[1]/td[5]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder65.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[4]/button/div/span[1]').text)
            over_under_x5_goals.append(driver_overunder65.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[2]/td[5]/button/div/span[1]').text)
            print(over_under_x5_goals)
        except:
            print_pro("[NO]* Over/Under 6.5 Goals")
            over_under_x5_goals+=["","","",""]

        if is_run == False:
            break

        #Correct Score
        try:
            correct_score_list = []
            for i in range(1,20):
                for j in range(4,6):
                    correct_score_list.append(driver_correctscore.find_element_by_xpath('//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr['+str(i)+']/td['+str(j)+']/button/div/span[1]').text)

            print(correct_score_list)
        except:
            print_pro("[NO]* Correct Score")
            correct_score_list = create_empty_list(38)

        if is_run == False:
            break

        #Both teams to Score?
        try:
            both_teams_to_score_list =[]
            for i in range(1, 3):
                for j in range(4, 6):
                    both_teams_to_score_list.append(driver_bothteamstoscore.find_element_by_xpath(
                        '//*[@id="main-wrapper"]/div/div[2]/div/ui-view/div/div/div[1]/div[3]/div/div[1]/div/bf-main-market/bf-main-marketview/div/div[2]/bf-marketview-runners-list[2]/div/div/div/table/tbody/tr[' + str(
                            i) + ']/td[' + str(j) + ']/button/div/span[1]').text)

            print(both_teams_to_score_list)
        except:
            print_pro("[NO]* Both teams to Score?")
            both_teams_to_score_list= create_empty_list(4)

        final_row = match_detials_list+half_time_list+first_half_goals_x5+half_time_score_list+match_odds_list+over_under_x5_goals+correct_score_list+both_teams_to_score_list
        wb = load_workbook(file_name+'.xlsx')
        ws = wb.active
        ws.append(final_row)
        wb.save(file_name+'.xlsx')
        print_pro("new data row added")

        if is_run == False:
            break
        print_pro("sleeping for "+str(sleep_time)+" seconds")
        time.sleep(sleep_time)
        print_pro("woke up")
        if is_run == False:
            break


window = tk.Tk()
window.geometry("700x360")
window.title("Football data scraper")
label_gameurl = tk.Label(text="Game url")
label_gameurl.place(x=10,y=10)
ent_gameurl = tk.Entry(width=102)
ent_gameurl.place(x=70,y=10)
ent_gameurl.bind('<Button-3>',rClicker, add='')

label_intervaltime = tk.Label(text="Time interval (in seconds)")
label_intervaltime.place(x=10,y=40)
ent_intervaltime = tk.Entry(width=5)
ent_intervaltime.place(x=160,y=40)

btn_start = tk.Button(text="Start", width=25, command=start_press)
btn_start.place(x=300,y=38)
btn_stop = tk.Button(text="Stop", width=25, command=stop_checking)
btn_stop.place(x=500,y=38)

label_date = tk.Label(text="", width=30, anchor='w')
label_date.place(x=10,y=70)
label_country = tk.Label(text="", width=30, anchor='w')
label_country.place(x=10,y=90)
label_series = tk.Label(text="", width=30, anchor='w')
label_series.place(x=10,y=110)
label_teams = tk.Label(text="", width=30, anchor='w')
label_teams.place(x=10,y=130)

label_score = tk.Label(text="", width=20, anchor='w')
label_score.place(x=300,y=90)
label_mins = tk.Label(text="", width=20, anchor='w')
label_mins.place(x=300,y=110)
btn_gameurl = tk.Button(text="Click here to open Betfair webpage ", width=30,command=open_match_page)
btn_gameurl.place(x=300,y=130)

text_logs = tk.Text(width=84, height=10)
text_logs.place(x=10, y=180)
window.mainloop()